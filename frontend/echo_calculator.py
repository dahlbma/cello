"""
Echo Spotting Calculator for Cello Application

This module generates Echo robot input files for compound dilution spotting.
It's integrated with the Cello application to use existing UI inputs and database connections.

Key Features:
- Uses data from dbInterface.getEchoData() instead of direct database queries
- Integrates with existing Cello UI inputs (userAddedVolume_eb, destinationPlates_eb, etc.)
- Generates Excel files compatible with Echo robot requirements
- Handles DMSO and control compound spotting with plate layout support
"""

import pandas as pd
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

class EchoSpotCalculator:
    """Echo spotting calculator for the Cello application"""
    
    # Constants
    DROPLET_SIZE_NL = 2.5  # Robot droplet size in nL
    DEST_PLATE_TYPE = "384PP_DMSO2"
    
    def __init__(self, logger_name="echo_calculator"):
        self.logger = logging.getLogger(logger_name)
        self.dmso_fixed_volume_nl = 250  # Default DMSO volume, can be overridden
        self.ctrl_fixed_volume_nl = 250  # Default CTRL volume, can be overridden
        self.dmso_source_index = 0  # Track which DMSO source well to use next (round-robin)
        self.max_dmso_percent = 0.01  # Default 1% limit, can be overridden from Excel
    
    def parse_excel_order(self, excel_file_path, max_dmso_pct_vol=None):
        """
        Parse the Excel order file with two sheets:
        Sheet 1: Order list (Compound Id, Batch Id, Final conc (nM))
        Sheet 2: Plate layout (384-well layout with DMSO/CTRL/Empty markings)
        
        Args:
            excel_file_path: Path to the Excel file
            max_dmso_pct_vol: Maximum DMSO percentage (e.g., 0.01 for 1%), 
                            typically from MAX_DMSO_pct_vol_eb QLineEdit
        
        Returns:
            tuple: (df_order, special_wells, compound_wells)
        """
        try:
            # Read both sheets
            df_order = pd.read_excel(excel_file_path, sheet_name=0)
            df_layout = pd.read_excel(excel_file_path, sheet_name=1)
            
            # Normalize headers
            df_order.columns = df_order.columns.str.strip()
            
            # Set MAX_DMSO_PERCENT from parameter
            if max_dmso_pct_vol is not None:
                try:
                    # Convert to float if it's a string (from QLineEdit.text())
                    max_dmso_value = float(max_dmso_pct_vol)
                    if max_dmso_value > 0:
                        self.max_dmso_percent = max_dmso_value
                        self.logger.info(f"MAX_DMSO_PERCENT set to {self.max_dmso_percent*100:.1f}% from parameter")
                    else:
                        self.logger.warning("MAX_DMSO_PERCENT parameter is <= 0, using default of 1%")
                except (ValueError, TypeError) as e:
                    self.logger.warning(f"Could not parse MAX_DMSO_PERCENT parameter: {e}, using default of 1%")
            else:
                self.logger.info(f"No MAX_DMSO_PERCENT parameter provided, using default of 1%")
            
            # Parse plate layout
            special_wells, compound_wells = self._parse_plate_layout(df_layout)
            
            return df_order, special_wells, compound_wells
            
        except Exception as e:
            self.logger.error(f"Error parsing Excel file: {e}")
            raise
    
    def _parse_plate_layout(self, df_layout):
        """
        Parse the plate layout from Sheet 2.
        Returns:
            tuple: (special_wells, compound_wells)
            - special_wells: List of tuples (well_position, content) for DMSO/CTRL wells
            - compound_wells: List of well_positions for available compound spots
        """
        special_wells = []
        compound_wells = []
        rows = "ABCDEFGHIJKLMNOP"
        
        for row_idx, row_data in df_layout.iterrows():
            row_letter = row_data['Unnamed: 0']  # Should be A, B, C, etc.
            if row_letter not in rows:
                continue
                
            # Iterate through columns 1-24
            for col_num in range(1, 25):
                if col_num in df_layout.columns:
                    content = row_data[col_num]
                    well_position = f"{row_letter}{col_num:02d}"
                    
                    # Skip explicitly marked as "Empty"
                    if pd.notna(content) and str(content).strip().upper() == 'EMPTY':
                        continue
                    
                    # Handle special wells (DMSO, CTRL)
                    if pd.notna(content):
                        content_str = str(content).strip()
                        if content_str.upper() == 'DMSO' or content_str.startswith('CTRL'):
                            special_wells.append((well_position, content_str))
                        else:
                            # Non-empty, non-DMSO, non-CTRL, non-Empty -> available for compounds
                            compound_wells.append(well_position)
                    else:
                        # NaN/blank cell -> available for compounds
                        compound_wells.append(well_position)
        
        return special_wells, compound_wells
    
    def parse_destination_plates(self, destination_plates_text):
        """
        Parse destination plates from the text field.
        Expected format: Whitespace-separated list of plate IDs (e.g., "P123456 P123457 P123458")
        
        Returns:
            list: List of destination plate IDs
        """
        if not destination_plates_text.strip():
            return []
        
        # Split by whitespace and filter out empty strings
        plates = [plate.strip() for plate in destination_plates_text.split() if plate.strip()]
        
        # Validate plate ID format (P followed by 6 digits)
        valid_plates = []
        for plate in plates:
            if len(plate) == 7 and plate.upper().startswith('P') and plate[1:].isdigit():
                valid_plates.append(plate.upper())
            else:
                self.logger.warning(f"Invalid plate ID format: {plate}")
        
        return valid_plates
    
    def calculate_transfer_volume(self, target_conc_nm, source_conc_mm, diluent_vol_ul):
        """
        Calculate exact transfer volume in nL required using dilution formula C1×V1 = C2×V2
        
        Args:
            target_conc_nm: Target concentration in nM
            source_conc_mm: Source concentration in mM
            diluent_vol_ul: Diluent volume in µL
            
        Returns:
            float: Required transfer volume in nL
        """
        if source_conc_mm <= 0:
            return 0
        return (target_conc_nm * diluent_vol_ul * 1000) / (source_conc_mm * 1_000_000)
    
    def snap_to_droplet_size(self, volume_nl):
        """
        Snap volume to nearest robot droplet size.
        
        Args:
            volume_nl: Volume in nL
            
        Returns:
            float: Snapped volume in nL (multiple of DROPLET_SIZE_NL)
        """
        steps = round(volume_nl / self.DROPLET_SIZE_NL)
        return steps * self.DROPLET_SIZE_NL
    
    def find_best_source(self, batch_id, target_nm, available_sources, diluent_vol_ul, max_conc_error_pct=15.0):
        """
        Find the best source for a given batch and target concentration.
        Strategy: Find the source that gives the MINIMUM concentration error within constraints.
        
        Args:
            batch_id: Batch ID to find
            target_nm: Target concentration in nM
            available_sources: List of source data dictionaries
            diluent_vol_ul: Diluent volume in µL
            max_conc_error_pct: Maximum acceptable concentration error percentage (default: 15%)
            
        Returns:
            tuple: (best_source_dict, transfer_volume_nl) or (None, 0) if not found
        """
        print(f"Finding best source for batch_id={batch_id}, target_nm={target_nm}, diluent_vol_ul={diluent_vol_ul}")
        # Filter sources by batch ID
        matches = [src for src in available_sources if src['batch_id'] == str(batch_id)]
        print(f"Found {matches} matching sources for batch_id={batch_id}")
        if not matches:
            return None, 0
        
        # Calculate max allowed transfer volume (DMSO limit from Excel or default)
        max_transfer_nl = diluent_vol_ul * 1000 * self.max_dmso_percent
        
        # Sort by concentration descending (highest first)
        matches_sorted = sorted(matches, key=lambda x: x['source_conc_mm'], reverse=True)
        
        best_candidate = None
        best_error = float('inf')
        best_within_tolerance = None
        best_error_within_tolerance = float('inf')
        
        for src in matches_sorted:
            # Calculate required volume
            req_vol = self.calculate_transfer_volume(target_nm, src['source_conc_mm'], diluent_vol_ul)
            
            # Snap to nearest droplet size
            snapped_vol = self.snap_to_droplet_size(req_vol)
            
            # Check constraints
            if snapped_vol <= 0:
                continue  # Source too concentrated
            
            if snapped_vol > max_transfer_nl:
                continue  # Would exceed DMSO limit
            
            # Calculate actual achieved concentration with snapped volume
            achieved_nm = (snapped_vol * src['source_conc_mm'] * 1_000_000) / (diluent_vol_ul * 1000)
            
            # Calculate concentration error percentage
            conc_error_pct = abs(achieved_nm - target_nm) / target_nm * 100
            
            # Track best candidate within tolerance
            if conc_error_pct <= max_conc_error_pct:
                if conc_error_pct < best_error_within_tolerance:
                    best_error_within_tolerance = conc_error_pct
                    best_within_tolerance = (src, snapped_vol)
            
            # Track overall best candidate (even if outside tolerance)
            if conc_error_pct < best_error:
                best_error = conc_error_pct
                best_candidate = (src, snapped_vol)
        
        # Return best candidate within tolerance if found
        if best_within_tolerance:
            return best_within_tolerance
        
        # Otherwise return best candidate if error < 50%
        if best_candidate and best_error < 50.0:
            return best_candidate
        
        return None, 0
    
    def _collect_dmso_sources(self, source_data, dmso_plate):
        """
        Collect all available DMSO source wells for round-robin distribution.
        
        Args:
            source_data: List of source plate data
            dmso_plate: DMSO plate ID
            
        Returns:
            list: List of DMSO source dictionaries
        """
        dmso_sources = []
        
        # First try: Look for DMSO compounds (by compound_id) from the specified DMSO plate
        if dmso_plate:
            dmso_sources = [src for src in source_data 
                          if src['source_plate'] == dmso_plate 
                          and src.get('compound_id', '').upper() == 'DMSO']
        
        # Second try: If no sources from dmso_plate, look for DMSO compound_id in any plate
        if not dmso_sources:
            dmso_sources = [src for src in source_data 
                          if src.get('compound_id', '').upper() == 'DMSO']
        
        # Third try: Look for any source that might be DMSO (batch_id contains DMSO)
        if not dmso_sources:
            dmso_sources = [src for src in source_data 
                          if 'DMSO' in str(src.get('batch_id', '')).upper()]
        
        return dmso_sources
    
    def _get_next_dmso_source(self):
        """
        Get the next DMSO source well using round-robin distribution.
        
        Returns:
            dict: DMSO source dictionary or None if no sources available
        """
        if not self.dmso_sources:
            return None
        
        # Get current source and increment index for next call
        dmso_src = self.dmso_sources[self.dmso_source_index]
        self.dmso_source_index = (self.dmso_source_index + 1) % len(self.dmso_sources)
        
        return dmso_src
    
    def generate_echo_file(self, order_data, source_data, ctrl_plate, dmso_plate, 
                          diluent_vol_ul, destination_plates, excel_order_path, output_path,
                          dmso_volume_nl=250, ctrl_volume_nl=250, backfill=False, well_order='horizontal', 
                          max_dmso_pct_vol=None, progress_callback=None):
        """
        Generate the Echo robot input file.
        
        Args:
            order_data: DataFrame with compound order (from Excel Sheet 1)
            source_data: List of source plate data from dbInterface.getEchoData()
            ctrl_plate: Control plate ID
            dmso_plate: DMSO plate ID
            diluent_vol_ul: Diluent volume in µL
            destination_plates: List of destination plate IDs
            excel_order_path: Path to Excel order file (for reading plate layout)
            output_path: Output file path
            dmso_volume_nl: Volume for DMSO transfers in nL (default: 250)
            ctrl_volume_nl: Volume for CTRL transfers in nL (default: 250)
            backfill: If True, backfill all non-DMSO wells with DMSO to equalize volumes
            well_order: 'horizontal' (A01, A02, ...) or 'vertical' (A01, B01, ...) (default: 'horizontal')
            max_dmso_pct_vol: Maximum DMSO percentage (e.g., 0.01 for 1%), typically from MAX_DMSO_pct_vol_eb QLineEdit
            progress_callback: Optional callback function(percent) for progress reporting
            
        Returns:
            bool: True if successful, False otherwise
        """
        # Set DMSO and CTRL volumes for this run
        self.dmso_fixed_volume_nl = dmso_volume_nl
        self.ctrl_fixed_volume_nl = ctrl_volume_nl
        self.dmso_source_index = 0  # Reset DMSO source index for each file generation
        
        try:
            # Parse plate layout from Excel file
            _, special_wells, compound_wells = self.parse_excel_order(excel_order_path, max_dmso_pct_vol)
            
            # Sort compound wells based on well_order preference
            if well_order.lower() == 'vertical':
                # Sort by column first (A01, B01, C01, ..., P01, A02, B02, ...)
                compound_wells = sorted(compound_wells, key=lambda w: (int(w[1:]), w[0]))
            else:
                # Sort by row first (A01, A02, ..., A24, B01, B02, ...) - this is default
                compound_wells = sorted(compound_wells, key=lambda w: (w[0], int(w[1:])))
            
            if not destination_plates:
                self.logger.error("No destination plates specified")
                return False
            
            self.logger.info(f"Parsed plate layout: {len(special_wells)} special wells, {len(compound_wells)} compound wells")
            
            output_rows = []
            max_transfer_nl = diluent_vol_ul * 1000 * self.max_dmso_percent
            
            self.logger.info(f"DMSO limit: Max transfer volume is {max_transfer_nl} nL per well ({self.max_dmso_percent*100:.1f}% of {diluent_vol_ul} µL)")
            
            # Collect all DMSO sources for round-robin distribution
            self.dmso_sources = self._collect_dmso_sources(source_data, dmso_plate)
            
            if self.dmso_sources:
                self.logger.info(f"Found {len(self.dmso_sources)} DMSO source wells for distribution:")
                for src in self.dmso_sources[:5]:  # Log first 5 examples
                    self.logger.info(f"  - Plate: {src['source_plate']}, Well: {src['source_well']}, Batch: {src.get('batch_id', 'N/A')}")
            else:
                self.logger.warning("No DMSO sources found in source data")
            
            # Process each destination plate
            compound_idx = 0
            total_compounds = len(order_data)
            total_wells = total_compounds + (len(special_wells) * len(destination_plates))
            wells_processed = 0
            
            for plate_num, dest_plate_id in enumerate(destination_plates, 1):
                # Add special wells (DMSO/CTRL) for this plate
                self._add_special_wells(output_rows, special_wells, dest_plate_id, 
                                      source_data, ctrl_plate, dmso_plate)
                wells_processed += len(special_wells)
                
                # Report progress
                if progress_callback:
                    progress_callback(int((wells_processed / total_wells) * 100))
                
                # Add compounds for this plate
                compounds_added = 0
                max_compounds_per_plate = len(compound_wells)
                
                while compounds_added < max_compounds_per_plate and compound_idx < total_compounds:
                    row = order_data.iloc[compound_idx]
                    dest_well = compound_wells[compounds_added]
                    
                    # Process this compound
                    success = self._process_compound(
                        output_rows, row, dest_well, dest_plate_id, 
                        source_data, diluent_vol_ul
                    )
                    
                    compound_idx += 1
                    compounds_added += 1
                    wells_processed += 1
                    
                    # Report progress periodically (every 10 compounds)
                    if progress_callback and compounds_added % 10 == 0:
                        progress_callback(int((wells_processed / total_wells) * 100))
                
                self.logger.info(f"Plate {plate_num}/{len(destination_plates)}: {compounds_added} compounds added")
                
                # Report progress after each plate
                if progress_callback:
                    progress_callback(int((wells_processed / total_wells) * 100))
                
                # If we've processed all compounds, break
                if compound_idx >= total_compounds:
                    break
            
            # Check if we processed all compounds
            if compound_idx < total_compounds:
                self.logger.warning(f"Not all compounds processed: {compound_idx}/{total_compounds}")
            
            # Apply backfill if requested - calculate and add inline
            if backfill:
                self.logger.info("Applying DMSO backfill to equalize well volumes")
                self._apply_backfill_inline(output_rows, source_data, dmso_plate)
            
            # Write output file
            if output_rows:
                if progress_callback:
                    progress_callback(95)
                self._write_excel_output(output_rows, output_path)
                if progress_callback:
                    progress_callback(100)
                self.logger.info(f"Echo file created: {output_path}")
                return True
            else:
                self.logger.error("No valid transfers generated")
                return False
                
        except Exception as e:
            self.logger.error(f"Error generating Echo file: {e}")
            return False
    
    def _add_special_wells(self, output_rows, special_wells, dest_plate_id, 
                          source_data, ctrl_plate, dmso_plate):
        """Add DMSO and control wells for a destination plate"""
        for dest_well, content in special_wells:
            if content.upper() == 'DMSO':
                # Get next DMSO source using round-robin distribution
                dmso_src = self._get_next_dmso_source()
                
                if dmso_src:
                    output_rows.append({
                        'Source plate name': dmso_src['source_plate'],
                        'Source Plate type': dmso_src['plate_subtype'],
                        'Source well': dmso_src['source_well'],
                        'Sample ID': 'DMSO',
                        'Sample name': dmso_src['batch_id'],
                        'Destination plate name': dest_plate_id,
                        'Destination Plate type': self.DEST_PLATE_TYPE,
                        'Destination well': dest_well,
                        'Transfer volym (nL)': self.dmso_fixed_volume_nl,
                        'Target conc (nM)': '',
                        'Final conc (nM)': '',
                        'Conc error (%)': '',
                        'Source conc (mM)': dmso_src['source_conc_mm'],
                        'Exception': ''
                    })
                else:
                    error_msg = f'DMSO source not found (searched plate: {dmso_plate}, total sources: {len(source_data)})'
                    self.logger.warning(error_msg)
                    output_rows.append({
                        'Source plate name': '',
                        'Source Plate type': '',
                        'Source well': '',
                        'Sample ID': 'DMSO',
                        'Sample name': '',
                        'Destination plate name': dest_plate_id,
                        'Destination Plate type': self.DEST_PLATE_TYPE,
                        'Destination well': dest_well,
                        'Transfer volym (nL)': self.dmso_fixed_volume_nl,
                        'Target conc (nM)': '',
                        'Final conc (nM)': '',
                        'Conc error (%)': '',
                        'Source conc (mM)': '',
                        'Exception': error_msg
                    })
            
            elif content.startswith('CTRL'):
                # Find control compound source - search across ALL source plates
                # (controls may be on list plates OR the ctrl_plate)
                ctrl_sources = [src for src in source_data 
                              if src.get('batch_id', '') == content]
                
                if ctrl_sources:
                    ctrl_src = ctrl_sources[0]
                    output_rows.append({
                        'Source plate name': ctrl_src['source_plate'],
                        'Source Plate type': ctrl_src['plate_subtype'],
                        'Source well': ctrl_src['source_well'],
                        'Sample ID': ctrl_src['compound_id'],
                        'Sample name': content,
                        'Destination plate name': dest_plate_id,
                        'Destination Plate type': self.DEST_PLATE_TYPE,
                        'Destination well': dest_well,
                        'Transfer volym (nL)': self.ctrl_fixed_volume_nl,
                        'Target conc (nM)': '',
                        'Final conc (nM)': '',
                        'Conc error (%)': '',
                        'Source conc (mM)': ctrl_src['source_conc_mm'],
                        'Exception': ''
                    })
                else:
                    output_rows.append({
                        'Source plate name': '',
                        'Source Plate type': '',
                        'Source well': '',
                        'Sample ID': 'CBK999999',
                        'Sample name': content,
                        'Destination plate name': dest_plate_id,
                        'Destination Plate type': self.DEST_PLATE_TYPE,
                        'Destination well': dest_well,
                        'Transfer volym (nL)': self.ctrl_fixed_volume_nl,
                        'Target conc (nM)': '',
                        'Final conc (nM)': '',
                        'Conc error (%)': '',
                        'Source conc (mM)': '',
                        'Exception': 'Control compound not found in database'
                    })
    
    def _process_compound(self, output_rows, row, dest_well, dest_plate_id, 
                         source_data, diluent_vol_ul):
        """Process a single compound entry"""
        batch_id = row['Batch Id']
        target_nm = row['Final conc (nM)']
        compound_id = row['Compound Id']
        
        # Find best source
        best_source, transfer_vol = self.find_best_source(
            batch_id, target_nm, source_data, diluent_vol_ul
        )
        
        if best_source and transfer_vol > 0:
            # Calculate actual achieved concentration with the snapped transfer volume
            achieved_nm = (transfer_vol * best_source['source_conc_mm'] * 1_000_000) / (diluent_vol_ul * 1000)
            
            # Calculate concentration error percentage
            conc_error_pct = ((achieved_nm - target_nm) / target_nm) * 100
            
            output_rows.append({
                'Source plate name': best_source['source_plate'],
                'Source Plate type': best_source['plate_subtype'],
                'Source well': best_source['source_well'],
                'Sample ID': compound_id,
                'Sample name': batch_id,
                'Destination plate name': dest_plate_id,
                'Destination Plate type': self.DEST_PLATE_TYPE,
                'Destination well': dest_well,
                'Transfer volym (nL)': transfer_vol,
                'Target conc (nM)': target_nm,
                'Final conc (nM)': achieved_nm,
                'Conc error (%)': conc_error_pct,
                'Source conc (mM)': best_source['source_conc_mm'],
                'Exception': ''
            })
            return True
        else:
            # Determine why it failed and provide helpful error message
            matches = [src for src in source_data if src['batch_id'] == batch_id]
            
            if not matches:
                error_msg = 'Compound not found in source plates'
            else:
                # Get available concentrations
                available_concs = sorted(set(src['source_conc_mm'] for src in matches), reverse=True)
                concs_str = ', '.join(f"{c:.3f}" for c in available_concs[:5])  # Show up to 5 concentrations
                if len(available_concs) > 5:
                    concs_str += f" (+ {len(available_concs) - 5} more)"
                
                # Calculate what concentration we could achieve with the best available source
                max_transfer_nl = diluent_vol_ul * 1000 * self.max_dmso_percent
                best_achievable_info = ""
                
                if available_concs:
                    best_conc_mm = available_concs[0]  # Highest concentration
                    req_vol = self.calculate_transfer_volume(target_nm, best_conc_mm, diluent_vol_ul)
                    snapped_vol = self.snap_to_droplet_size(req_vol)
                    
                    if snapped_vol > max_transfer_nl:
                        # Source too dilute - calculate what we could achieve with max volume
                        achievable_nm = (best_conc_mm * 1_000_000 * max_transfer_nl) / (diluent_vol_ul * 1000)
                        best_achievable_info = f", best achievable: {achievable_nm:.1f} nM using {best_conc_mm:.3f} mM"
                    elif snapped_vol <= 0:
                        # Source too concentrated - would need less than one droplet
                        best_achievable_info = f", {best_conc_mm:.3f} mM is too concentrated"
                
                error_msg = f'No valid source conc found (available: {concs_str} mM{best_achievable_info})'
            
            # Add failed entry
            output_rows.append({
                'Source plate name': '',
                'Source Plate type': '',
                'Source well': '',
                'Sample ID': compound_id,
                'Sample name': batch_id,
                'Destination plate name': dest_plate_id,
                'Destination Plate type': self.DEST_PLATE_TYPE,
                'Destination well': dest_well,
                'Transfer volym (nL)': '',
                'Target conc (nM)': target_nm,
                'Final conc (nM)': '',
                'Conc error (%)': '',
                'Source conc (mM)': '',
                'Exception': error_msg
            })
            self.logger.warning(f"Failed to find valid source for {batch_id} at {target_nm} nM: {error_msg}")
            return False
    
    def _apply_backfill_inline(self, output_rows, source_data, dmso_plate):
        """
        Apply DMSO backfill to equalize volumes across all non-DMSO wells.
        Inserts backfill rows immediately after the last transfer for each well.
        Uses round-robin distribution across DMSO source wells.
        
        Args:
            output_rows: List of output row dictionaries (modified in place)
            source_data: List of source plate data
            dmso_plate: DMSO plate ID for finding DMSO source
        """
        # Check if DMSO sources are available
        if not self.dmso_sources:
            self.logger.error("Cannot apply backfill: No DMSO sources available")
            return
        
        self.logger.info(f"Using {len(self.dmso_sources)} DMSO source wells for backfill distribution")
        
        # Track volumes and last index for each well
        well_info = {}  # Key: (dest_plate, dest_well), Value: {'volume': total, 'last_idx': index, 'is_ctrl': bool}
        
        for idx, row in enumerate(output_rows):
            dest_plate = row['Destination plate name']
            dest_well = row['Destination well']
            transfer_vol = row['Transfer volym (nL)']
            sample_id = row['Sample ID']
            sample_name = row['Sample name']
            
            # Skip if this is already a DMSO-only well or has no volume
            if sample_id == 'DMSO':
                continue
            
            if not transfer_vol or str(transfer_vol).strip() == '':
                continue
            
            try:
                volume = float(transfer_vol)
            except (ValueError, TypeError):
                continue
            
            # Check if this is a CTRL well
            is_ctrl = str(sample_name).startswith('CTRL')
            
            key = (dest_plate, dest_well)
            if key not in well_info:
                well_info[key] = {'volume': 0, 'last_idx': idx, 'is_ctrl': is_ctrl}
            
            well_info[key]['volume'] += volume
            well_info[key]['last_idx'] = idx  # Track last row for this well
            # If any transfer to this well is a CTRL, mark it as such
            if is_ctrl:
                well_info[key]['is_ctrl'] = True
        
        if not well_info:
            self.logger.warning("No non-DMSO wells found for backfill")
            return
        
        # Find maximum volume among non-CTRL wells only
        non_ctrl_volumes = [info['volume'] for info in well_info.values() if not info['is_ctrl']]
        
        if not non_ctrl_volumes:
            self.logger.warning("No non-CTRL compound wells found for backfill")
            return
        
        max_volume = max(non_ctrl_volumes)
        self.logger.info(f"Maximum non-DMSO/non-CTRL well volume: {max_volume} nL")
        
        # Create backfill rows with their insertion indices
        backfill_inserts = []  # List of (insert_after_idx, backfill_row)
        
        for (dest_plate, dest_well), info in well_info.items():
            # Skip CTRL wells - they should not be backfilled
            if info['is_ctrl']:
                continue
            
            current_volume = info['volume']
            if current_volume < max_volume:
                backfill_volume = max_volume - current_volume
                
                # Snap to droplet size
                backfill_volume = self.snap_to_droplet_size(backfill_volume)
                
                if backfill_volume > 0:
                    # Get next DMSO source using round-robin distribution
                    dmso_src = self._get_next_dmso_source()
                    
                    if dmso_src:
                        backfill_row = {
                            'Source plate name': dmso_src['source_plate'],
                            'Source Plate type': dmso_src['plate_subtype'],
                            'Source well': dmso_src['source_well'],
                            'Sample ID': 'BACKFILL',
                            'Sample name': 'BACKFILL',
                            'Destination plate name': dest_plate,
                            'Destination Plate type': self.DEST_PLATE_TYPE,
                            'Destination well': dest_well,
                            'Transfer volym (nL)': backfill_volume,
                            'Target conc (nM)': '',
                            'Final conc (nM)': '',
                            'Conc error (%)': '',
                            'Source conc (mM)': dmso_src.get('source_conc_mm', ''),
                            'Exception': ''
                        }
                        backfill_inserts.append((info['last_idx'], backfill_row))
        
        # Sort by index in reverse order so we can insert from back to front
        # (this way indices remain valid as we insert)
        backfill_inserts.sort(key=lambda x: x[0], reverse=True)
        
        # Insert backfill rows
        for insert_after_idx, backfill_row in backfill_inserts:
            output_rows.insert(insert_after_idx + 1, backfill_row)
        
        self.logger.info(f"Added {len(backfill_inserts)} DMSO backfill transfers inline")
    
    def _write_excel_output(self, output_rows, output_path):
        """Write the output data to Excel file with proper formatting"""
        df_out = pd.DataFrame(output_rows)
        
        # Ensure correct column order
        cols = ['Source plate name', 'Source Plate type', 'Source well', 'Sample ID',
                'Sample name', 'Destination plate name', 'Destination Plate type',
                'Destination well', 'Transfer volym (nL)', 'Target conc (nM)', 'Final conc (nM)',
                'Conc error (%)', 'Source conc (mM)', 'Exception']
        df_out = df_out[cols]
        
        # Round numeric columns to 3 decimal places for clarity
        numeric_cols = ['Transfer volym (nL)', 'Target conc (nM)', 'Final conc (nM)', 
                       'Conc error (%)', 'Source conc (mM)']
        for col in numeric_cols:
            # Only round if column has numeric data
            df_out[col] = pd.to_numeric(df_out[col], errors='coerce').round(3)
        
        # Write with Excel formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_out.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # Yellow fill for failed transfers
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            # Orange fill for high concentration errors (warning)
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            
            # Auto-adjust column widths
            for i, col in enumerate(df_out.columns):
                # Calculate max length
                max_len = len(str(col))
                if not df_out[col].empty:
                    max_data_len = df_out[col].astype(str).map(len).max()
                    if max_data_len > max_len:
                        max_len = max_data_len
                
                # Set column width with padding
                adjusted_width = min(max_len + 2, 50)  # Cap at 50 characters
                col_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Highlight rows with exceptions (yellow) or high concentration errors (orange)
            exception_col_idx = cols.index('Exception') + 1
            conc_error_col_idx = cols.index('Conc error (%)') + 1
            
            for row_idx in range(2, len(df_out) + 2):  # Skip header
                # First check for exceptions (yellow - highest priority)
                exception_value = worksheet.cell(row=row_idx, column=exception_col_idx).value
                if exception_value and str(exception_value).strip() and str(exception_value).lower() != 'nan':
                    for col_idx in range(1, len(cols) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = yellow_fill
                # Otherwise check for high concentration error (orange)
                else:
                    conc_error_value = worksheet.cell(row=row_idx, column=conc_error_col_idx).value
                    if conc_error_value and str(conc_error_value).strip() and str(conc_error_value).lower() != 'nan':
                        try:
                            error_pct = abs(float(conc_error_value))
                            if error_pct > 15.0:  # 15% threshold for warning
                                for col_idx in range(1, len(cols) + 1):
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    cell.fill = orange_fill
                        except (ValueError, TypeError):
                            pass  # Skip if cannot parse